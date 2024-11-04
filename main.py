import os
import openpyxl
from xml.etree import ElementTree as ET
from xml.dom import minidom
from tkinter import filedialog, messagebox
import tkinter as tk
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import re
import zipfile


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Global variable for output file path
output_file_path = None
target_language = None


# Function to style the Excel sheet
def style_excel_sheet(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color="D3D3D3"),
                         right=Side(style='thin', color="D3D3D3"),
                         top=Side(style='thin', color="D3D3D3"),
                         bottom=Side(style='thin', color="D3D3D3"))

    for col in ws.columns:
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:  # Apply header style
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    ws.sheet_view.showGridLines = False


# Refactored function to convert XLIFF to Excel without saving
def xliff_to_excel(xliff_file):
    try:
        tree = ET.parse(xliff_file)
        root = tree.getroot()

        # Extract target-language value
        target_language = root.find(".//file").attrib.get("target-language", "translations")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = target_language  # Set the worksheet title to the target-language

        # Create headers
        ws.append(["ID", "Max Width", "Size Unit", "Source", "Target", "Note"])

        # Add data to the sheet
        for file_element in root.findall("file"):
            for trans_unit in file_element.find("body").findall("trans-unit"):
                id_value = trans_unit.get("id", "")
                max_width = trans_unit.get("maxwidth", "")
                size_unit = trans_unit.get("size-unit", "")
                source_text = trans_unit.find("source").text if trans_unit.find("source") is not None else ""
                target_text = trans_unit.find("target").text if trans_unit.find("target") is not None else ""
                note_text = trans_unit.find("note").text if trans_unit.find("note") is not None else ""

                # Cast max_width to an integer or float if it's a number
                try:
                    max_width = int(max_width) if max_width.isdigit() else float(max_width)
                except (ValueError, TypeError):
                    max_width = None  # If it's not a valid number, keep it as None

                ws.append([id_value, max_width, size_unit, source_text, target_text, note_text])

        # Apply styling to the worksheet
        style_excel_sheet(ws)

        return wb, target_language

    except Exception as e:
        logging.error(f"An error occurred during the XLIFF to Excel conversion: {e}")
        print(f"An error occurred during the XLIFF to Excel conversion: {e}")
        raise



# Function to convert Excel to XLIFF
def excel_to_xliff(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        logging.info("Excel workbook loaded successfully.")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            root = ET.Element("xliff", version="1.2")
            file_attributes = {
                "original": "Salesforce",
                "source-language": "en_US",
                "target-language": sheet_name,
                "translation-type": "metadata",
                "datatype": "xml"
            }
            file_element = ET.SubElement(root, "file", **file_attributes)
            body_element = ET.SubElement(file_element, "body")

            seen_ids = set()  # To keep track of seen Ids and avoid duplicates
            for row in sheet.iter_rows(min_row=2, values_only=True):
                id_value = str(row[0])  # Assuming the ID is in the first column
                if id_value not in seen_ids:  # Check if the ID is already processed
                    seen_ids.add(id_value)  # Mark this ID as seen
                    trans_unit_element = ET.SubElement(body_element, "trans-unit", id=id_value, maxwidth=str(row[1]),
                                                       size_unit=str(row[2]))
                    source_element = ET.SubElement(trans_unit_element, "source")
                    source_element.text = f"{str(row[3])}"
                    target_element = ET.SubElement(trans_unit_element, "target")
                    target_text = str(row[4]) if row[4] else "<>"  # Set "<>" if Target is None or empty
                    target_element.text = target_text

                    if len(row) > 5 and row[5]:
                        note_element = ET.SubElement(trans_unit_element, "note")
                        note_element.text = f"{str(row[5])}"

            xml_declaration = '<?xml version="1.0" encoding="UTF-8"?>\n'
            xml_string = minidom.parseString(ET.tostring(root)).toprettyxml(indent="    ")
            xml_string = xml_string.split('\n', 1)[1] if xml_string.startswith('<?xml') else xml_string

            full_xml = xml_declaration + xml_string

            output_file_path = filedialog.asksaveasfilename(
                title="Save File As",
                defaultextension=".xlf",
                filetypes=[("XLIFF files", "*.xlf"), ("All Files", "*.*")],
                initialfile=f"{sheet_name}_output.xlf"
            )

            if output_file_path:
                with open(output_file_path, "w", encoding="utf-8") as f:
                    f.write(full_xml)
                logging.info(f"File saved successfully at {output_file_path}")
                print(f"File saved successfully at {output_file_path}")
            else:
                logging.info("File save operation was cancelled.")
                print("File save operation was cancelled.")

        logging.info("Conversion complete.")
        print("Conversion complete.")
    except Exception as e:
        logging.error(f"An error occurred during the conversion: {e}")
        print(f"An error occurred during the conversion: {e}")
        raise


# Function to select an Excel file and convert it to XLIFF
def select_excel_to_xliff():
    excel_file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx;*.xls"), ]
    )
    if excel_file_path:
        logging.info("Excel file selected.")
        print("Excel file selected.")
        excel_to_xliff(excel_file_path)
    else:
        logging.info("No Excel file selected. Exiting.")
        print("No Excel file selected. Exiting.")


# Function to handle single XLIFF to Excel conversion
def select_xliff_to_excel():
    xliff_file_path = filedialog.askopenfilename(
        title="Select XLIFF File",
        filetypes=[("XLIFF files", "*.xlf"), ("All Files", "*.*")]
    )
    if xliff_file_path:
        logging.info("XLIFF file selected.")
        print("XLIFF file selected.")

        try:
            wb, target_language = xliff_to_excel(xliff_file_path)

            # Set default output file name using target-language
            default_output_filename = f"Excel to xlf {target_language}.xlsx"
            output_file_path = filedialog.asksaveasfilename(
                title="Save Excel File As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=default_output_filename
            )

            if output_file_path:
                wb.save(output_file_path)
                logging.info(f"XLIFF converted to Excel and saved at {output_file_path}")
                print(f"XLIFF converted to Excel and saved at {output_file_path}")
            else:
                logging.info("File save operation was cancelled.")
                print("File save operation was cancelled.")

        except Exception as e:
            logging.error(f"An error occurred during the XLIFF to Excel conversion: {e}")
            print(f"An error occurred during the XLIFF to Excel conversion: {e}")
            raise

    else:
        logging.info("No XLIFF file selected. Exiting.")
        print("No XLIFF file selected. Exiting.")


# New function to handle multiple XLIFF to Excel conversion using the refactored function
def multiple_xliff_to_excel():
    xliff_files = filedialog.askopenfilenames(
        title="Select Multiple XLIFF Files",
        filetypes=[("XLIFF files", "*.xlf"), ("All Files", "*.*")]
    )

    if not xliff_files:
        logging.info("No files selected. Exiting.")
        return

    # Ask the user to select a base folder for saving the Excel files
    base_folder = filedialog.askdirectory(title="Select Base Folder for Output")
    if not base_folder:
        logging.info("No base folder selected. Exiting.")
        return

    for xliff_file in xliff_files:
        try:
            # Use the refactored xliff_to_excel function to get the workbook and language
            wb, target_language = xliff_to_excel(xliff_file)

            # Create a folder for the target language if it doesn't exist
            folder_path = os.path.join(base_folder, target_language)
            os.makedirs(folder_path, exist_ok=True)

            # Define the output file path for the Excel file in the respective folder
            excel_filename = f"Excel to xlf {target_language}.xlsx"
            output_file_path = os.path.join(folder_path, excel_filename)

            # Save the Excel file
            wb.save(output_file_path)
            logging.info(f"XLIFF converted to Excel and saved at {output_file_path}")
            print(f"XLIFF converted to Excel and saved at {output_file_path}")

        except Exception as e:
            logging.error(f"An error occurred while processing {xliff_file}: {e}")
            print(f"An error occurred while processing {xliff_file}: {e}")

    logging.info("Multiple files processed successfully.")
    messagebox.showinfo("Success", "All files have been processed and saved.")

def select_two_files(root):
    # Create a new window for selecting files
    select_window = tk.Toplevel(root)
    select_window.title("Select Source Language and Optional English Files")
    select_window.geometry("400x200")

    source_file_path = None
    english_file_path = None

    # Function to select the source language file
    def select_source_file():
        nonlocal source_file_path

        source_file_path = filedialog.askopenfilename(
            title="Select Source Language File",
            filetypes=[("XLIFF files", "*.xlf"), ("All Files", "*.*")]
        )
        if source_file_path:
            source_file_label.config(text=f"Source File: {source_file_path}")
            logging.info(f"Source language file selected: {source_file_path}")
        else:
            logging.info("No source language file selected.")

    # Function to select the English file (optional)
    def select_english_file():
        nonlocal english_file_path
        english_file_path = filedialog.askopenfilename(
            title="Select English File (Optional)",
            filetypes=[("XLIFF files", "*.xlf"), ("All Files", "*.*")]
        )
        if english_file_path:
            english_file_label.config(text=f"English File: {english_file_path}")
            logging.info(f"English file selected: {english_file_path}")
        else:
            logging.info("No English file selected. English translations will not be added.")

    # Function to process the files after selection
    def process_files():
        global output_file_path  # Reuse the global output_file_path

        if not source_file_path:
            messagebox.showerror("Error", "Please select the source language file.")
            return

        try:
            # Convert source XLIFF to Excel and save it
            logging.info("Converting source language file to Excel...")
            wb, target_language = xliff_to_excel(source_file_path)

            # Prompt the user to save the Excel file
            output_file_path = filedialog.asksaveasfilename(
                title="Save Source Language Excel File As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=f"{target_language}_source.xlsx"
            )

            if output_file_path:
                wb.save(output_file_path)
                logging.info(f"Source Excel file saved at {output_file_path}")
            else:
                logging.info("File save operation was cancelled.")
                return  # Exit if the user cancels the save operation

            # Optional: Process the English file to extract Target values, if available
            english_translation_map = {}
            if english_file_path:
                logging.info("Extracting Target values from the English file...")
                tree = ET.parse(english_file_path)
                root = tree.getroot()

                for file_element in root.findall("file"):
                    for trans_unit in file_element.find("body").findall("trans-unit"):
                        id_value = trans_unit.get("id", "")
                        target_text = trans_unit.find("target").text if trans_unit.find("target") is not None else ""
                        english_translation_map[id_value] = target_text

            # Load the previously saved source Excel file
            logging.info(f"Loading the previously saved Excel file from {output_file_path}...")
            wb = openpyxl.load_workbook(output_file_path)
            ws = wb.active

            # If the English file is selected, add "Translated to English" column
            if english_file_path:
                # Add columns "Feedback By Customer" and "Feedback for Length"
                ws["H1"] = "Feedback By Customer"
                ws["I1"] = "Feedback for Length"
                ws["G1"] = "Translated to English"

                for row in ws.iter_rows(min_row=2, max_col=6):  # Assuming data goes from column A to F
                    id_value = str(row[0].value)  # Column A contains IDs
                    english_translation = english_translation_map.get(id_value, "")
                    ws[f"G{row[0].row}"] = english_translation  # Insert the English translation in column G

                    # Insert the formula for "Feedback for Length" (Column I)
                    feedback_formula = (f'=IF(H{row[0].row}="","",IF(LEN(H{row[0].row})>B{row[0].row},"* The new '
                                        f'translation is too long ("&LEN(H{row[0].row})&") should be under '
                                        f'"&B{row[0].row}&" chars","OK"))')
                    ws[f"I{row[0].row}"] = feedback_formula
            else:
                ws["G1"] = "Feedback By Customer"
                ws["H1"] = "Feedback for Length"
                # Iterate over the rows of the source file and match IDs with English translations if available
                for row in ws.iter_rows(min_row=2, max_col=6):  # Assuming data goes from column A to F
                    feedback_formula = (f'=IF(G{row[0].row}="","",IF(LEN(G{row[0].row})>B{row[0].row},"* The new '
                                        f'translation is too long ("&LEN(G{row[0].row})&") should be under "'
                                        f'&B{row[0].row}&" chars","OK"))')
                    ws[f"H{row[0].row}"] = feedback_formula

            # Apply styling to the entire sheet, including the new columns
            style_excel_sheet(ws)

            # Save the modified Excel file
            output_file_path = filedialog.asksaveasfilename(
                title="Save Modified Excel File As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=f"{target_language}_with_Feedback.xlsx"
            )

            if output_file_path:
                wb.save(output_file_path)
                logging.info(f"Source file updated with feedback saved at {output_file_path}")
                messagebox.showinfo("Success", f"File saved successfully at {output_file_path}")
            else:
                logging.info("File save operation was cancelled.")
                print("File save operation was cancelled.")

        except Exception as e:
            logging.error(f"An error occurred during the processing: {e}")
            messagebox.showerror("Error", f"An error occurred: {e}")

    # Buttons for selecting files
    source_file_btn = tk.Button(select_window, text="Select Source Language File", command=select_source_file)
    source_file_btn.pack(pady=10)

    english_file_btn = tk.Button(select_window, text="Select English File (Optional)", command=select_english_file)
    english_file_btn.pack(pady=10)

    # Labels to display selected file paths
    source_file_label = tk.Label(select_window, text="Source File: Not selected")
    source_file_label.pack(pady=5)

    english_file_label = tk.Label(select_window, text="English File (Optional): Not selected")
    english_file_label.pack(pady=5)

    # Button to start processing
    process_btn = tk.Button(select_window, text="Process Files", command=process_files)
    process_btn.pack(pady=10)

    # Close button to close the window
    close_btn = tk.Button(select_window, text="Close", command=select_window.destroy)
    close_btn.pack(pady=10)


def create_package(root):
    # Hide the root window (if not already hidden)
    root.withdraw()

    # Ask the user to select multiple .objectTranslation files
    input_file_paths = filedialog.askopenfilenames(title="Select .objectTranslation files",
                                                   filetypes=[("Object Translation Files", "*.objectTranslation")])

    # Ask the user to select the base folder to save zipped deployment packages
    base_output_folder = filedialog.askdirectory(title="Select the base folder to save deployment packages")

    # Define sections to remove
    sections_to_remove = ['fields', 'validationRules', 'webLinks', 'layouts', 'fieldSets']

    # Dictionary to keep track of files for each language
    language_files = {}

    # Process each selected file
    if input_file_paths and base_output_folder:
        for input_file_path in input_file_paths:
            # Read the input file content
            with open(input_file_path, 'r', encoding='utf-8') as file:
                file_content = file.read()

            # Remove all specified sections and any resulting blank lines
            for section in sections_to_remove:
                file_content = re.sub(rf'<{section}>.*?</{section}>\s*', '', file_content, flags=re.DOTALL)

            # Extract the object API name and language code from the filename
            file_name = os.path.basename(input_file_path)
            object_api_name, language_code = file_name.split('-')
            language_code = language_code.replace('.objectTranslation', '')

            # Organize files by language in a dictionary
            if language_code not in language_files:
                language_files[language_code] = set()
            language_files[language_code].add(object_api_name)

            # Define the language-specific "unpackaged/translations" folder
            unpackaged_folder = os.path.join(base_output_folder, language_code, "unpackaged", "objectTranslations")
            os.makedirs(unpackaged_folder, exist_ok=True)  # Create the folder if it doesn't exist

            # Save the modified .objectTranslation file in the translations folder
            modified_file_path = os.path.join(unpackaged_folder,
                                              file_name.replace('.objectTranslation', '.objectTranslation'))
            with open(modified_file_path, 'w', encoding='utf-8') as modified_file:
                modified_file.write(file_content)

        # For each language, create a package.xml and zip the contents within an "unpackaged" folder
        for language_code, object_api_names in language_files.items():
            # Generate package.xml for the current language
            package_xml_content = '''<?xml version="1.0" encoding="UTF-8"?>
    <Package xmlns="http://soap.sforce.com/2006/04/metadata">
        <types>
    '''
            for api_name in sorted(object_api_names):
                # Include the language code in each member entry as <API_NAME>-<language_code>
                package_xml_content += f'        <members>{api_name}-{language_code}</members>\n'

            package_xml_content += '''        <name>CustomObjectTranslation</name>
        </types>
        <version>57.0</version>
    </Package>'''

            # Save the package.xml in the "unpackaged" folder for the language
            package_xml_path = os.path.join(base_output_folder, language_code, "unpackaged", "package.xml")
            with open(package_xml_path, 'w', encoding='utf-8') as package_file:
                package_file.write(package_xml_content)

            # Zip the "unpackaged" folder for deployment
            unpackaged_folder_path = os.path.join(base_output_folder, language_code, "unpackaged")
            zip_file_path = os.path.join(base_output_folder, f"{language_code}_deployment_package.zip")
            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root_dir, _, files in os.walk(unpackaged_folder_path):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        zipf.write(file_path, os.path.relpath(file_path, os.path.join(base_output_folder, language_code)))

            print(f"Deployment package for language '{language_code}' created: {zip_file_path}")
    else:
        print("File selection was cancelled.")


# GUI update to add the new button
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Excel to XLIFF Converter")
    root.geometry("300x260")

    btn_excel_to_xliff = tk.Button(root, text="Excel to XLIFF", command=select_excel_to_xliff, width=20)
    btn_excel_to_xliff.pack(pady=10)

    btn_xliff_to_excel = tk.Button(root, text="XLIFF to Excel", command=select_xliff_to_excel, width=20)
    btn_xliff_to_excel.pack(pady=10)

    btn_multiple_xliff_to_excel = tk.Button(root, text="Multiple Files XLIFF to Excel", command=multiple_xliff_to_excel,
                                            width=20)
    btn_multiple_xliff_to_excel.pack(pady=10)

    btn_select_files = tk.Button(root, text="Feedback file automation", command=lambda: select_two_files(root),
                                 width=20)
    btn_select_files.pack(pady=10)

    btn_create_package = tk.Button(root, text="Create Package (Tabs and Labels)", command=lambda: create_package(root)
                                   , width=20)
    btn_create_package.pack(pady=10)

    lbl_version = tk.Label(root, text="Version 1.1")
    lbl_version.pack(pady=10)

    root.mainloop()
